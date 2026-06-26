"""SL2: real-file KB ingestion tests for CSV, XLSX, PDF, MD and TXT."""

from __future__ import annotations

import io
from typing import Any

import pytest
from fastapi.testclient import TestClient
from fpdf import FPDF
from openpyxl import Workbook


@pytest.fixture()
def client(tmp_path: Any, monkeypatch: pytest.MonkeyPatch) -> TestClient:
    db_path = tmp_path / "spaceloom.sqlite3"
    audit_path = tmp_path / "audit.jsonl"
    monkeypatch.setenv("SPACELOOM_DB_PATH", str(db_path))

    for name in (
        "OPENAI_API_KEY",
        "SPACELOOM_OPENAI_API_KEY",
        "ANTHROPIC_API_KEY",
        "SPACELOOM_ANTHROPIC_API_KEY",
        "GOOGLE_API_KEY",
        "GEMINI_API_KEY",
        "SPACELOOM_GOOGLE_API_KEY",
        "SPACELOOM_ENABLE_OLLAMA",
        "SPACELOOM_OLLAMA_ENABLED",
        "SPACELOOM_PROVIDER_ALLOWLIST",
        "SPACELOOM_BUDGET_CAP_USD",
        "SPACELOOM_DEV_TRUST_HEADERS",
    ):
        monkeypatch.delenv(name, raising=False)

    from app.src.audit import audit_writer
    from app.src.main import create_app

    audit_writer.audit_path = audit_path
    with TestClient(create_app()) as test_client:
        yield test_client


def _demo_workspace_id(client: TestClient) -> str:
    response = client.get("/api/workspaces")
    assert response.status_code == 200
    workspaces = response.json()["workspaces"]
    assert workspaces
    return workspaces[0]["id"]


def _create_workspace(client: TestClient, name: str) -> str:
    response = client.post("/api/workspaces", json={"name": name})
    assert response.status_code == 201, response.text
    return response.json()["id"]


def _upload_file(
    client: TestClient,
    workspace_id: str,
    filename: str,
    content_bytes: bytes,
    mime_type: str,
    title: str = "Upload",
) -> dict[str, Any]:
    response = client.post(
        f"/api/workspaces/{workspace_id}/kb/upload",
        data={"title": title, "source_version": "v1"},
        files={"file": (filename, io.BytesIO(content_bytes), mime_type)},
    )
    assert response.status_code == 201, response.text
    return response.json()


def _make_xlsx_bytes() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Catalogo"
    sheet.append(["sku", "precio_usd", "moneda", "stock", "vigente_desde", "vigente_hasta"])
    sheet.append(["TEL-001", 12.50, "USD", 240, "2026-01-01", "2026-12-31"])
    sheet.append(["TEL-002", 18.00, "USD", 120, "2026-01-01", "2026-12-31"])
    buf = io.BytesIO()
    workbook.save(buf)
    return buf.getvalue()


def _make_pdf_bytes() -> bytes:
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", size=12)
    pdf.cell(text="Catalogo de telas")
    pdf.ln()
    with pdf.table() as table:
        row = table.row()
        row.cell("sku")
        row.cell("precio_usd")
        row.cell("stock")
        row = table.row()
        row.cell("TEL-001")
        row.cell("12.50")
        row.cell("240")
    return pdf.output()


def test_upload_csv_creates_facts_and_chunks(client: TestClient) -> None:
    workspace_id = _demo_workspace_id(client)
    csv_text = "sku,nombre,precio_usd,moneda,stock,vigente_desde,vigente_hasta\nTEL-001,Oxford,12.50,USD,240,2026-01-01,2026-12-31"
    source = _upload_file(
        client,
        workspace_id,
        "catalogo.csv",
        csv_text.encode("utf-8"),
        "text/csv",
        title="Catalogo CSV",
    )
    assert source["type"] == "csv"
    assert source["file_name"] == "catalogo.csv"
    assert source["mime_type"] == "text/csv"
    assert source["file_size"] > 0
    assert source["parser_version"] == "builtin"

    response = client.get(f"/api/workspaces/{workspace_id}/kb/search?q=TEL-001")
    assert response.status_code == 200
    data = response.json()
    assert any(fact["entity_key"] == "TEL-001" for fact in data["facts"])
    assert any("12.50" in fact["field_value"] for fact in data["facts"])


def test_upload_xlsx_extracts_facts_with_source_sheet(client: TestClient) -> None:
    workspace_id = _demo_workspace_id(client)
    xlsx_bytes = _make_xlsx_bytes()
    source = _upload_file(
        client,
        workspace_id,
        "catalogo.xlsx",
        xlsx_bytes,
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        title="Catalogo XLSX",
    )
    assert source["type"] == "xlsx"
    assert source["parser_version"].startswith("openpyxl")

    response = client.get(f"/api/workspaces/{workspace_id}/kb/search?q=TEL-001")
    assert response.status_code == 200
    data = response.json()
    facts = data["facts"]
    assert any(fact["entity_key"] == "TEL-001" for fact in facts)
    sheet_facts = [fact for fact in facts if fact.get("source_sheet")]
    assert all(fact["source_sheet"] == "Catalogo" for fact in sheet_facts)


def test_upload_pdf_extracts_text_and_tables(client: TestClient) -> None:
    workspace_id = _demo_workspace_id(client)
    pdf_bytes = _make_pdf_bytes()
    source = _upload_file(
        client,
        workspace_id,
        "catalogo.pdf",
        pdf_bytes,
        "application/pdf",
        title="Catalogo PDF",
    )
    assert source["type"] == "pdf"
    assert source["parser_version"].startswith("pdfplumber")

    response = client.get(f"/api/workspaces/{workspace_id}/kb/search?q=TEL-001")
    assert response.status_code == 200
    data = response.json()
    assert any("Catalogo" in chunk["content_text"] for chunk in data["chunks"])
    assert any(fact["entity_key"] == "TEL-001" for fact in data["facts"])


def test_upload_md_and_txt(client: TestClient) -> None:
    workspace_id = _demo_workspace_id(client)
    md_bytes = b"# Pricing\n\nOxford USD 12.50 por metro.\n"
    md = _upload_file(client, workspace_id, "pricing.md", md_bytes, "text/markdown", title="Pricing MD")
    assert md["type"] == "md"
    assert md["parser_version"] == "builtin"

    txt_bytes = b"Lino USD 18.00 por metro.\n"
    txt = _upload_file(client, workspace_id, "pricing.txt", txt_bytes, "text/plain", title="Pricing TXT")
    assert txt["type"] == "txt"
    assert txt["parser_version"] == "builtin"

    response = client.get(f"/api/workspaces/{workspace_id}/kb/search?q=Lino")
    assert response.status_code == 200
    assert any("Lino" in chunk["content_text"] for chunk in response.json()["chunks"])


def test_upload_rejects_unsupported_extension(client: TestClient) -> None:
    workspace_id = _demo_workspace_id(client)
    response = client.post(
        f"/api/workspaces/{workspace_id}/kb/upload",
        data={"title": "Bad"},
        files={"file": ("page.html", io.BytesIO(b"<html></html>"), "text/html")},
    )
    assert response.status_code == 415


def test_upload_rejects_csv_formula_injection(client: TestClient) -> None:
    workspace_id = _demo_workspace_id(client)
    malicious_csv = "sku,price\n=cmd|'/C calc'!A1,12.50"
    response = client.post(
        f"/api/workspaces/{workspace_id}/kb/upload",
        data={"title": "Bad CSV"},
        files={"file": ("bad.csv", io.BytesIO(malicious_csv.encode("utf-8")), "text/csv")},
    )
    assert response.status_code == 422


def test_upload_is_workspace_isolated(client: TestClient) -> None:
    ws_a = _create_workspace(client, "Workspace A")
    ws_b = _create_workspace(client, "Workspace B")
    csv_text = "sku,price\nTEL-001,12.50"
    source_a = _upload_file(client, ws_a, "a.csv", csv_text.encode("utf-8"), "text/csv")

    response_b = client.get(f"/api/workspaces/{ws_b}/kb/sources")
    assert response_b.status_code == 200
    assert source_a["id"] not in {s["id"] for s in response_b.json()}

    response_a = client.get(f"/api/workspaces/{ws_a}/kb/sources")
    assert source_a["id"] in {s["id"] for s in response_a.json()}
