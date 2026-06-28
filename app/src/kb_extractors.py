"""FaberLoom SL2: extractors for PDF and XLSX KB sources.

Extractors turn binary files into plain text and structured tables so the core
KB pipeline can chunk and fact-ify them uniformly.  Extraction is deliberately
read-only: we never execute macros, scripts or formulas.
"""

from __future__ import annotations

import csv
import io
import re
import zipfile
from dataclasses import dataclass
from typing import Any

import chardet
import filetype
import openpyxl
import pdfplumber
from pdfminer.pdfdocument import PDFDocument
from pdfminer.pdfparser import PDFParser
from pdfminer.pdftypes import resolve1
from pdfminer.psparser import PSLiteral


MAX_EXTRACTED_CHARS = 500_000


@dataclass
class ExtractedTable:
    """A table extracted from a file, ready to become kb_fact rows."""

    name: str | None = None
    rows: list[dict[str, str]] | None = None
    headers: list[str] | None = None


@dataclass
class ExtractedDocument:
    """Result of extracting a binary KB source."""

    text: str = ""
    tables: list[ExtractedTable] | None = None
    file_name: str | None = None
    mime_type: str | None = None
    warnings: list[str] | None = None


def _clean_cell(value: Any) -> str:
    """Convert an extracted cell value to a safe string."""

    if value is None:
        return ""
    text = str(value).strip()
    # Collapse whitespace and newlines.
    text = re.sub(r"\s+", " ", text)
    return text


def _sanitize_table_rows(rows: list[dict[str, str]]) -> list[dict[str, str]]:
    """Clean keys and values of an extracted table."""

    cleaned: list[dict[str, str]] = []
    for row in rows:
        clean_row = {}
        for key, value in row.items():
            clean_key = _clean_cell(key)
            clean_value = _clean_cell(value)
            if clean_key:
                clean_row[clean_key] = clean_value
        if any(v.strip() for v in clean_row.values()):
            cleaned.append(clean_row)
    return cleaned


def _pdf_contains_javascript(file_bytes: bytes) -> bool:
    """Return True if the PDF catalog declares JavaScript actions."""

    def _name_str(value: Any) -> str:
        if isinstance(value, PSLiteral):
            return value.name
        return str(value)

    def _has_js(value: Any) -> bool:
        try:
            resolved = resolve1(value)
        except Exception:
            return False
        if isinstance(resolved, dict):
            if _name_str(resolved.get("S")) == "JavaScript":
                return True
            return any(_has_js(v) for v in resolved.values())
        if isinstance(resolved, list):
            return any(_has_js(v) for v in resolved)
        return False

    try:
        parser = PDFParser(io.BytesIO(file_bytes))
        doc = PDFDocument(parser)
        if _has_js(doc.catalog):
            return True
        pages = resolve1(doc.catalog.get("Pages"))
        if pages and _has_js(pages):
            return True
    except Exception:
        # If we cannot parse the PDF, let the normal extractor report it.
        return False
    return False


def extract_from_pdf(file_bytes: bytes, file_name: str | None = None) -> ExtractedDocument:
    """Extract text and tables from a PDF.

    Raises ValueError for password-protected or unreadable PDFs.
    """

    if _pdf_contains_javascript(file_bytes):
        raise ValueError("PDF contains JavaScript actions and is not allowed as a KB source.")

    warnings: list[str] = []
    text_parts: list[str] = []
    tables: list[ExtractedTable] = []

    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page_index, page in enumerate(pdf.pages, start=1):
                page_text = page.extract_text() or ""
                if page_text:
                    text_parts.append(page_text)

                page_tables = page.extract_tables()
                for table_index, table in enumerate(page_tables, start=1):
                    if not table or len(table) < 2:
                        continue
                    headers = [_clean_cell(h) for h in table[0]]
                    rows: list[dict[str, str]] = []
                    for row in table[1:]:
                        row_dict = {}
                        for col_index, header in enumerate(headers):
                            if header:
                                value = row[col_index] if col_index < len(row) else ""
                                row_dict[header] = _clean_cell(value)
                        if any(v.strip() for v in row_dict.values()):
                            rows.append(row_dict)
                    if rows:
                        tables.append(
                            ExtractedTable(
                                name=f"page {page_index} table {table_index}",
                                rows=_sanitize_table_rows(rows),
                                headers=headers,
                            )
                        )
    except pdfplumber.utils.exceptions.PDFException as exc:
        raise ValueError(f"Could not parse PDF: {exc}") from exc
    except Exception as exc:
        raise ValueError(f"PDF extraction failed: {exc}") from exc

    full_text = "\n\n".join(text_parts).strip()
    if not full_text and not any(t.rows for t in tables):
        warnings.append("PDF appears to contain no selectable text or tables.")

    return ExtractedDocument(
        text=full_text[:MAX_EXTRACTED_CHARS],
        tables=tables or None,
        file_name=file_name,
        mime_type="application/pdf",
        warnings=warnings or None,
    )


def _xlsx_contains_macros(file_bytes: bytes) -> bool:
    """Return True if an OOXML workbook contains a VBA project."""

    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            return any(name.lower().startswith("xl/") and name.lower().endswith("vbaproject.bin") for name in zf.namelist())
    except zipfile.BadZipFile:
        return False


def extract_from_xlsx(file_bytes: bytes, file_name: str | None = None) -> ExtractedDocument:
    """Extract text and tables from an Excel workbook.

    Raises ValueError for password-protected, corrupt, or macro-enabled workbooks.
    """

    if _xlsx_contains_macros(file_bytes):
        raise ValueError("XLSX contains macros and is not allowed as a KB source.")

    warnings: list[str] = []
    text_parts: list[str] = []
    tables: list[ExtractedTable] = []

    try:
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_bytes),
            data_only=True,
            read_only=True,
        )
    except Exception as exc:
        raise ValueError(f"Could not parse XLSX: {exc}") from exc

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        sheet_rows: list[list[Any]] = []
        for row in sheet.iter_rows(values_only=True):
            sheet_rows.append(list(row))

        if not sheet_rows:
            continue

        # Detect header row: first row with any non-empty cell.
        headers = [_clean_cell(cell) for cell in sheet_rows[0]]
        data_rows = sheet_rows[1:]

        rows: list[dict[str, str]] = []
        for row in data_rows:
            row_dict = {}
            for col_index, header in enumerate(headers):
                if header and col_index < len(row):
                    row_dict[header] = _clean_cell(row[col_index])
            if any(v.strip() for v in row_dict.values()):
                rows.append(row_dict)

        if rows:
            tables.append(
                ExtractedTable(
                    name=sheet_name,
                    rows=_sanitize_table_rows(rows),
                    headers=headers,
                )
            )
            # Also append a textual representation for chunking.
            text_parts.append(
                f"## Sheet: {sheet_name}\n" + "\n".join(
                    ", ".join(f"{k}: {v}" for k, v in row.items())
                    for row in rows
                )
            )

    workbook.close()

    full_text = "\n\n".join(text_parts).strip()
    if not full_text and not any(t.rows for t in tables):
        warnings.append("Workbook appears to be empty.")

    return ExtractedDocument(
        text=full_text[:MAX_EXTRACTED_CHARS],
        tables=tables or None,
        file_name=file_name,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        warnings=warnings or None,
    )


def extract_from_csv_text(content_text: str) -> ExtractedDocument:
    """Parse CSV text into an ExtractedDocument with a single table."""

    try:
        reader = csv.DictReader(io.StringIO(content_text))
        rows = list(reader)
    except Exception as exc:
        raise ValueError(f"Invalid CSV: {exc}") from exc

    if not rows:
        return ExtractedDocument(text=content_text)

    headers = list(rows[0].keys())
    sanitized = _sanitize_table_rows(rows)
    return ExtractedDocument(
        text=content_text[:MAX_EXTRACTED_CHARS],
        tables=[ExtractedTable(name="csv", rows=sanitized, headers=headers)],
        mime_type="text/csv",
    )


def decode_text_file(file_bytes: bytes, file_name: str | None = None) -> ExtractedDocument:
    """Decode a plain text / markdown file, guessing encoding."""

    detected = chardet.detect(file_bytes)
    encoding = detected.get("encoding") or "utf-8"
    confidence = detected.get("confidence") or 0.0
    warnings: list[str] | None = None
    if confidence < 0.5:
        warnings = [f"Low-confidence encoding detection ({confidence:.2f}); falling back to utf-8."]
        encoding = "utf-8"

    try:
        text = file_bytes.decode(encoding)
    except (UnicodeDecodeError, LookupError):
        text = file_bytes.decode("utf-8", errors="replace")
        warnings = ["File decoded with replacement characters; some content may be lost."]

    return ExtractedDocument(
        text=text[:MAX_EXTRACTED_CHARS],
        file_name=file_name,
        mime_type="text/plain",
        warnings=warnings,
    )


def detect_mime_type(file_bytes: bytes, file_name: str | None = None) -> str:
    """Return a MIME type from file bytes and optional name."""

    kind = filetype.guess(file_bytes)
    if kind:
        return kind.mime
    if file_name:
        name_lower = file_name.lower()
        if name_lower.endswith(".pdf"):
            return "application/pdf"
        if name_lower.endswith(".xlsx"):
            return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        if name_lower.endswith(".csv"):
            return "text/csv"
        if name_lower.endswith(".md"):
            return "text/markdown"
        if name_lower.endswith(".txt"):
            return "text/plain"
    return "application/octet-stream"


def extract_document(
    file_bytes: bytes,
    *,
    file_name: str | None = None,
    mime_type: str | None = None,
) -> ExtractedDocument:
    """Dispatch to the right extractor based on MIME type."""

    mime = mime_type or detect_mime_type(file_bytes, file_name)

    if mime == "application/pdf":
        return extract_from_pdf(file_bytes, file_name)
    if mime in {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    }:
        return extract_from_xlsx(file_bytes, file_name)
    if mime == "text/csv":
        return extract_from_csv_text(file_bytes.decode("utf-8", errors="replace"))
    if mime in {"text/markdown", "text/plain"}:
        return decode_text_file(file_bytes, file_name)

    # Fallback: try plain text decoding.
    return decode_text_file(file_bytes, file_name)
